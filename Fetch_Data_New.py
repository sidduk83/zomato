import re
from bs4 import BeautifulSoup
import urllib
from urllib.parse import urlparse
from selenium import webdriver
from bs4 import NavigableString
import json
import pandas as pd
from openpyxl import load_workbook

filepath = r"C:\Users\Zprint\Desktop\Learning\Tableau\Zomato\Zomato_Fetch_Py.xlsx"

#rest_details = {'name': '8293', 'rating': '3.5', 'area': 'KR Puram', 'type': 'Quick Bites', 'reviews': '31', 'cuisines': 'Chinese', 'cost_for_two': '₹300', 'geo_location': ['13.0102343470', '77.7064486966']}

book = load_workbook(filepath)
writer = pd.ExcelWriter(filepath, engine='openpyxl')
writer.book = book
ws = book['Sheet1']
row_no = 2
browser = webdriver.Firefox(executable_path=r'C:\Users\Zprint\Downloads\geckodriver.exe')
for i in range(2,1678):
    if ws.cell(row=i, column=1).value:
        
        browser.get(ws.cell(row=i, column=1).value)

        html_text = browser.page_source
        soup = BeautifulSoup(html_text, 'html.parser')

        rest_details = dict()
        name_anchor = soup.find("a", attrs={"class": "ui large header left"})
        if name_anchor:
            rest_details['name'] = name_anchor.text.strip()
        else:
            rest_details['name'] = ''

        rating_div = soup.find("div", attrs={"class": re.compile("rating-for")})
        if rating_div:
            rest_details['rating'] = rating_div.text.strip()[:-2]
        else:
            rest_details['rating'] = 'N/A'  # default

        area_div = soup.find("a",attrs={"class": re.compile("left grey-text fontsize3")})
        if area_div:
            rest_details['area'] = area_div.text.strip()
        else:
            rest_details['area'] = 'N/A'  # default


        text_data = []
        for txt in soup.find_all("a",{"class" : ["grey-text fontsize3","View all"]}):
            text_data.append(txt.text.strip())
        text_data.sort()
        type_div = ','.join(text_data)
        ws.cell(row=i, column=5).value= type_div

        reviews_div = soup.find("span",attrs={"class": re.compile("tooltip_formatted fbold")})
        if reviews_div:
            rest_details['reviews'] = int(''.join(filter(str.isdigit, reviews_div.text.strip())))
        else:
            rest_details['reviews'] = 'N/A'  # default

        text_data = []
        for txt in soup.find_all(["div"],{"class" : "res-info-cuisines clearfix"}):
            text_data.append(txt.text.strip())
        text_data.sort()
        cuisines_div = ','.join(text_data)
        ws.cell(row=i, column=7).value= cuisines_div

        cost_for_two_div = soup.find("span",attrs={"aria-label": re.compile("for two people")})
        if cost_for_two_div:
            rest_details['cost_for_two'] =  int(''.join(filter(str.isdigit, cost_for_two_div.text.strip())))
        else:
            rest_details['cost_for_two'] = 'N/A'  # default

        geo_locale = soup.find("div", attrs={"class": "resmap-img"})
        if geo_locale:
            geo_url = geo_locale.attrs['data-url']
            parsed_url = urlparse(geo_url)
            geo_arr = str(urllib.parse.parse_qs(parsed_url.query)['center']).split(',')
            rest_details['geo_location'] = [re.sub("[^0-9\.]", "", geo_arr[0]), re.sub("[^0-9\.]", "", geo_arr[1])]
        if 'geo_location' not in rest_details:
            rest_details['geo_location'] = ['undefined', 'undefined']

        more_info_div = soup.find("div",attrs={"class": "res-info-feature-text"})
        if more_info_div:
            rest_details['more_info'] = more_info_div.text.strip()
        else:
            rest_details['more_info'] = 'N/A'  # default

        text_data = []
        for txt in soup.find_all(["span"],{"itemprop": "paymentAccepted"}):
            text_data.append(txt.text.strip())
        text_data.sort()
        payment_div = ','.join(text_data)
        ws.cell(row=i, column=11).value= payment_div

        ws.cell(row=i, column=2).value= rest_details["name"]
        try:
            ws.cell(row=i, column=3).value= float(rest_details["rating"])
        except:
            ws.cell(row=i, column=3).value= rest_details["rating"]
        ws.cell(row=i, column=4).value= rest_details["area"]
        #ws.cell(row=i, column=5).value= rest_details["type"]
        try:
            ws.cell(row=i, column=6).value= int(rest_details["reviews"])
        except:
            ws.cell(row=i, column=6).value= rest_details["reviews"]
        #ws.cell(row=i, column=7).value= rest_details["cuisines"]
        try:
            ws.cell(row=i, column=8).value= int(rest_details["cost_for_two"])
        except:
            ws.cell(row=i, column=8).value= rest_details["cost_for_two"]
        ws.cell(row=i, column=9).value= rest_details["geo_location"][0] + "," + rest_details["geo_location"][1]
        #ws.cell(row=i, column=11).value= rest_details["payment"]

        try:
            table = soup.find_all('table')[0]
            rows = table.findAll('tr')
            text_data = []
            for tr in rows:
                cols = tr.findAll('td',attrs={"class":"pl10"})
                for td in cols:
                    text = '-'.join(td)               
                    text_data.append(text.replace('Noon','PM'))
            text_data.sort()
            timings = ','.join(text_data)
            ws.cell(row=i, column=12).value= timings
        except:
            pass

        #rows = soup.find_all('res-info-feature-text')
        text_data = []
        
        for txt in soup.find_all('div',class_="res-info-feature-text"):
            text_data.append(txt.text.strip())
        text_data.sort()
        more_info = ','.join(text_data)
        ws.cell(row=i, column=10).value= more_info

        text_data = []
        text_filter = ()

        divTag = soup.find_all("div", ["grey-text"])
        for tag in divTag:
            if tag.text.strip() == "Food":
                classTag = soup.find_all("div", {"rv_highlights__score_bar mt5 mb5",re.compile("block level")})
##                for tag2 in classTag:
##                    print(rest_details["name"])
##                    print(tag2)
                    
            #text_data.append(int(''.join(filter(str.isdigit, str(tag)))[-1]))
        #print(text_data)
        
        writer.save()
browser.close()
